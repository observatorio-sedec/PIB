import datetime
from pathlib import Path
import pandas as pd
import openpyxl
import requests as rq
import ssl
from ajustar_planilha import ajustar_bordas, ajustar_colunas
ROOT_PATH = Path(__file__).parent


class TLSAdapter(rq.adapters.HTTPAdapter):
    def init_poolmanager(self, *args, **kwargs):
        ctx = ssl.create_default_context()
        ctx.set_ciphers("DEFAULT@SECLEVEL=1")
        ctx.options |= 0x4   # OP_LEGACY_SERVER_CONNECT
        kwargs["ssl_context"] = ctx
        return super(TLSAdapter, self).init_poolmanager(*args, **kwargs)

def requisitando_dados(api):
    with rq.session() as s:
        s.mount("https://", TLSAdapter())
        dados_brutos_api = s.get(api, verify=True)

    if dados_brutos_api.status_code != 200:
        raise Exception(f"A solicitação à API falhou com o código de status: {dados_brutos_api.status_code}")

    try:
        dados_brutos = dados_brutos_api.json()
    except Exception as e:
        raise Exception(f"Erro ao analisar a resposta JSON da API: {str(e)}")
    return dados_brutos

def extrair_dados(api, tabela_id):
    dados_brutos = requisitando_dados(api)

    variaveis_por_tabela = {
        5938: ['variavel37', 'variavel543', 'variavel498', 'variavel513', 'variavel517', 'variavel6575', 'variavel525'],
        5939: ['variavel529', 'variavel531', 'variavel532', 'variavel6568', 'variavel534']
    }

    if tabela_id in variaveis_por_tabela:
        variaveis = variaveis_por_tabela[tabela_id]

        if dados_brutos:
            return tuple(dados_brutos[i] if i < len(dados_brutos) else None for i in range(len(variaveis)))
        else:
            return tuple([None] * len(variaveis))
    
    return None

    
def tratando_dados5938(variavel37, variavel543, variavel498, variavel513, variavel517, variavel6575, variavel525):
    dados_limpos_37 = []
    dados_limpos_543 = []
    dados_limpos_498 = []
    dados_limpos_513 = []
    dados_limpos_517 = []
    dados_limpos_6575 = []
    dados_limpos_525 = []

    variaveis = [variavel37, variavel543, variavel498, variavel513, variavel517, variavel6575, variavel525]
    for i in variaveis:
        id_tabela = i['id']
        variavel = i['variavel']
        unidade = i['unidade']
        dados = i['resultados']
    
        for ii in dados:
            dados_produto = ii['classificacoes']
            dados_producao = ii['series']

            for iv in dados_producao:
                id = iv['localidade']['id']
                local = iv['localidade']['nome']
                dados_ano_producao = iv['serie']
                

                for ano, producao in dados_ano_producao.items():
                    producao = producao.replace('-', '0').replace('...', '0')
                    
                    
                    dict = {
                        'id': id,
                        'local': local,
                        #'id_produto': id_produto,
                        #'Categoria': nome_produto,
                        variavel: producao,
                        'unidade': unidade,
                        'ano': f'01/01/{ano}'
                }
                    if id_tabela == '37':
                        dados_limpos_37.append(dict)
                    elif id_tabela == '543':
                        dados_limpos_543.append(dict)
                    elif id_tabela == '498':
                        dados_limpos_498.append(dict)
                    elif id_tabela == '513':
                        dados_limpos_513.append(dict)
                    elif id_tabela == '517':
                        dados_limpos_517.append(dict)
                    elif id_tabela == '6575':
                        dados_limpos_6575.append(dict)
                    elif id_tabela == '525':
                        dados_limpos_525.append(dict)
                                
    return dados_limpos_37, dados_limpos_543, dados_limpos_498, dados_limpos_513, dados_limpos_517, dados_limpos_6575, dados_limpos_525

def gerando_dataframe5938(dados_limpos_37, dados_limpos_543, dados_limpos_498, dados_limpos_513, dados_limpos_517, dados_limpos_6575, dados_limpos_525):
    df37 = pd.DataFrame(dados_limpos_37)
    df543 = pd.DataFrame(dados_limpos_543)
    df498 = pd.DataFrame(dados_limpos_498)
    df513 = pd.DataFrame(dados_limpos_513)
    df517 = pd.DataFrame(dados_limpos_517)
    df6575 = pd.DataFrame(dados_limpos_6575)
    df525 = pd.DataFrame(dados_limpos_525)
    dataframe = pd.merge(df37, df543, on=['id', 'local', 'unidade', 'ano'], how='inner')
    dataframe = pd.merge(dataframe, df498, on=['id', 'local', 'unidade', 'ano'], how='inner')
    dataframe = pd.merge(dataframe, df513, on=['id', 'local', 'unidade', 'ano'], how='inner')
    dataframe = pd.merge(dataframe, df517, on=['id', 'local', 'unidade', 'ano'], how='inner')
    dataframe = pd.merge(dataframe, df6575, on=['id', 'local', 'unidade', 'ano'], how='inner')
    dataframe = pd.merge(dataframe, df525, on=['id', 'local', 'unidade', 'ano'], how='inner')
    
    dataframe[['Produto Interno Bruto a preços correntes', 'Impostos, líquidos de subsídios, sobre produtos a preços correntes', 'Valor adicionado bruto a preços correntes total', 'Valor adicionado bruto a preços correntes da agropecuária', 'Valor adicionado bruto a preços correntes da indústria', 'Valor adicionado bruto a preços correntes dos serviços, exclusive administração, defesa, educação e saúde públicas e seguridade social', 'Valor adicionado bruto a preços correntes da administração, defesa, educação e saúde públicas e seguridade social']] = \
    dataframe[['Produto Interno Bruto a preços correntes', 'Impostos, líquidos de subsídios, sobre produtos a preços correntes', 'Valor adicionado bruto a preços correntes total', 'Valor adicionado bruto a preços correntes da agropecuária', 'Valor adicionado bruto a preços correntes da indústria', 'Valor adicionado bruto a preços correntes dos serviços, exclusive administração, defesa, educação e saúde públicas e seguridade social', 'Valor adicionado bruto a preços correntes da administração, defesa, educação e saúde públicas e seguridade social']].astype(float)
    dataframe[['Produto Interno Bruto a preços correntes', 'Impostos, líquidos de subsídios, sobre produtos a preços correntes', 'Valor adicionado bruto a preços correntes total', 'Valor adicionado bruto a preços correntes da agropecuária', 'Valor adicionado bruto a preços correntes da indústria', 'Valor adicionado bruto a preços correntes dos serviços, exclusive administração, defesa, educação e saúde públicas e seguridade social', 'Valor adicionado bruto a preços correntes da administração, defesa, educação e saúde públicas e seguridade social']] = \
    dataframe[['Produto Interno Bruto a preços correntes', 'Impostos, líquidos de subsídios, sobre produtos a preços correntes', 'Valor adicionado bruto a preços correntes total', 'Valor adicionado bruto a preços correntes da agropecuária', 'Valor adicionado bruto a preços correntes da indústria', 'Valor adicionado bruto a preços correntes dos serviços, exclusive administração, defesa, educação e saúde públicas e seguridade social', 'Valor adicionado bruto a preços correntes da administração, defesa, educação e saúde públicas e seguridade social']] * 1000
    dataframe['local'] = dataframe['local'].str.replace(r'\s*\(MT\)', '', regex=True)
    return dataframe

def tratando_dados5939(variavel529, variavel531, variavel532, variavel6568, variavel534):
    dados_limpos_529 = []
    dados_limpos_531 = []
    dados_limpos_532 = []
    dados_limpos_6568 = []
    dados_limpos_534 = []


    variaveis = [variavel529, variavel531, variavel532, variavel6568, variavel534 ]
    for i in variaveis:
        id_tabela = i['id']
        variavel = i['variavel']
        unidade = i['unidade']
        dados = i['resultados']
    
        for ii in dados:
            dados_produto = ii['classificacoes']
            dados_producao = ii['series']

            for iv in dados_producao:
                id = iv['localidade']['id']
                local = iv['localidade']['nome']
                dados_ano_producao = iv['serie']
                

                for ano, producao in dados_ano_producao.items():
                    producao = producao.replace('-', '0').replace('...', '0')
                    
                    
                    dict = {
                        'id': id,
                        'local': local,
                        'ano': f'01/01/{ano}',
                        #'id_produto': id_produto,
                        #'Categoria': nome_produto,
                        variavel: producao,
                        'unidade': unidade
                }
                    if id_tabela == '529':
                        dados_limpos_529.append(dict)
                    elif id_tabela == '531':
                        dados_limpos_531.append(dict)
                    elif id_tabela == '532':
                        dados_limpos_532.append(dict)
                    elif id_tabela == '6568':
                        dados_limpos_6568.append(dict)
                    elif id_tabela == '534':
                        dados_limpos_534.append(dict)
                                
    return dados_limpos_529, dados_limpos_531, dados_limpos_532, dados_limpos_6568, dados_limpos_534

def gerando_dataframe5939(dados_limpos_529, dados_limpos_531, dados_limpos_532, dados_limpos_6568, dados_limpos_534):
    df529 = pd.DataFrame(dados_limpos_529)
    df531 = pd.DataFrame(dados_limpos_531)
    df532 = pd.DataFrame(dados_limpos_532)
    df6568 = pd.DataFrame(dados_limpos_6568)
    df534 = pd.DataFrame(dados_limpos_534)
    
    dataframe = pd.merge(df529, df531, on=['id', 'local', 'unidade', 'ano'], how='inner')
    dataframe = pd.merge(dataframe, df532, on=['id', 'local', 'unidade', 'ano'], how='inner')
    dataframe = pd.merge(dataframe, df6568, on=['id', 'local', 'unidade', 'ano'], how='inner')
    dataframe = pd.merge(dataframe, df534, on=['id', 'local', 'unidade', 'ano'], how='inner')
    dataframe[['Índice de Gini da distribuição do produto interno bruto a preços correntes', 'Índice de Gini da distribuição do valor adicionado bruto a preços correntes da agropecuária', 'Índice de Gini da distribuição do valor adicionado bruto a preços correntes da indústria', 'Índice de Gini da distribuição do valor adicionado bruto a preços correntes dos serviços, exclusive administração, defesa, educação e saúde públicas e seguridade social', 'Índice de Gini da distribuição do valor adicionado bruto a preços correntes da administração, defesa, educação e saúde públicas e seguridade social']] = \
    dataframe[['Índice de Gini da distribuição do produto interno bruto a preços correntes', 'Índice de Gini da distribuição do valor adicionado bruto a preços correntes da agropecuária', 'Índice de Gini da distribuição do valor adicionado bruto a preços correntes da indústria', 'Índice de Gini da distribuição do valor adicionado bruto a preços correntes dos serviços, exclusive administração, defesa, educação e saúde públicas e seguridade social', 'Índice de Gini da distribuição do valor adicionado bruto a preços correntes da administração, defesa, educação e saúde públicas e seguridade social']].astype(float)
    
    dataframe = dataframe.rename(columns={'ano': 'data'})
    return dataframe

ano_atual = datetime.datetime.now().year

def executando_dados(tabela, muni):
    lista_dados_37 = [] 
    lista_dados_543 = []
    lista_dados_498 = []
    lista_dados_513 = []
    lista_dados_517 = []
    lista_dados_6575 = []
    lista_dados_525 = []
    for ano in range(2002, ano_atual):
        if muni == True:
            api = f'https://servicodados.ibge.gov.br/api/v3/agregados/5938/periodos/{ano}/variaveis/37|543|498|513|517|6575|525?localidades=N6[5100102,5100201,5100250,5100300,5100359,5100409,5100508,5100607,5100805,5101001,5101209,5101258,5101308,5101407,5101605,5101704,5101803,5101852,5101902,5102504,5102603,5102637,5102678,5102686,5102694,5102702,5102793,5102850,5103007,5103056,5103106,5103205,5103254,5103304,5103353,5103361,5103379,5103403,5103437,5103452,5103502,5103601,5103700,5103809,5103858,5103908,5103957,5104104,5104203,5104500,5104526,5104542,5104559,5104609,5104807,5104906,5105002,5105101,5105150,5105176,5105200,5105234,5105259,5105309,5105507,5105580,5105606,5105622,5105903,5106000,5106109,5106158,5106174,5106182,5106190,5106208,5106216,5106224,5106232,5106240,5106257,5106265,5106273,5106281,5106299,5106307,5106315,5106372,5106422,5106455,5106505,5106653,5106703,5106752,5106778,5106802,5106828,5106851,5107008,5107040,5107065,5107107,5107156,5107180,5107198,5107206,5107248,5107263,5107297,5107305,5107354,5107404,5107578,5107602,5107701,5107743,5107750,5107768,5107776,5107792,5107800,5107859,5107875,5107883,5107909,5107925,5107941,5107958,5108006,5108055,5108105,5108204,5108303,5108352,5108402,5108501,5108600,5108808,5108857,5108907,5108956]'
        else:
            api = f'https://servicodados.ibge.gov.br/api/v3/agregados/5938/periodos/{ano}/variaveis/37|543|498|513|517|6575|525?localidades=N3[all]'
        variavel37, variavel543, variavel498, variavel513, variavel517, variavel6575, variavel525 = extrair_dados(api, tabela)
        if all(v is None for v in [variavel37, variavel543, variavel498, variavel513, variavel517, variavel6575, variavel525]):
            break
        else:
            novos_dados_37, novos_dados_543, novos_dados_498, novos_dados_513, novos_dados_517, novos_dados_6575, novos_dados_525 = tratando_dados5938(variavel37, variavel543, variavel498, variavel513, variavel517, variavel6575, variavel525)
            
            lista_dados_37.extend(novos_dados_37)
            lista_dados_543.extend(novos_dados_543)
            lista_dados_498.extend(novos_dados_498)
            lista_dados_513.extend(novos_dados_513)
            lista_dados_517.extend(novos_dados_517)
            lista_dados_6575.extend(novos_dados_6575)
            lista_dados_525.extend(novos_dados_525)
    
    return  lista_dados_37,lista_dados_543,lista_dados_498, lista_dados_513, lista_dados_517, lista_dados_6575, lista_dados_525

def executando_dados_2(tabela):
    lista_dados_529 = [] 
    lista_dados_531 = []
    lista_dados_532 = []
    lista_dados_6568 = []
    lista_dados_534 = []
    for ano in range(2002, ano_atual):
        api = f'https://servicodados.ibge.gov.br/api/v3/agregados/5939/periodos/{ano}/variaveis/529|531|532|6568|534?localidades=N3[all]'
        variavel529, variavel531, variavel532, variavel6568, variavel534 = extrair_dados(api, tabela)
        if all(v is None for v in [variavel529, variavel531, variavel532, variavel6568, variavel534]):
            break
        else:
            novos_dados_529, novos_dados_531, novos_dados_532, novos_dados_6568, novos_dados_534= tratando_dados5939(variavel529, variavel531, variavel532, variavel6568, variavel534)
            
            lista_dados_529.extend(novos_dados_529)
            lista_dados_531.extend(novos_dados_531)
            lista_dados_532.extend(novos_dados_532)
            lista_dados_6568.extend(novos_dados_6568)
            lista_dados_534.extend(novos_dados_534)
    
    return  lista_dados_529,lista_dados_531,lista_dados_532, lista_dados_6568, lista_dados_534


dados_limpos_37, dados_limpos_543, dados_limpos_498, dados_limpos_513, dados_limpos_517, dados_limpos_6575, dados_limpos_525 = executando_dados(5938, True)
df_pib_muni = gerando_dataframe5938(dados_limpos_37, dados_limpos_543, dados_limpos_498, dados_limpos_513, dados_limpos_517, dados_limpos_6575, dados_limpos_525)
df_pib_muni.to_excel(ROOT_PATH / 'dados_PIB_municipais.xlsx')

dados_limpos_37_estadual, dados_limpos_543_estadual, dados_limpos_498_estadual, dados_limpos_513_estadual, dados_limpos_517_estadual, dados_limpos_6575_estadual, dados_limpos_525_estadual = executando_dados(5938, False)
df_pib_estadual = gerando_dataframe5938(dados_limpos_37_estadual, dados_limpos_543_estadual, dados_limpos_498_estadual, dados_limpos_513_estadual, dados_limpos_517_estadual, dados_limpos_6575_estadual, dados_limpos_525_estadual)
df_pib_estadual.to_excel(ROOT_PATH / 'dados_PIB_estaduais.xlsx')

dados_limpos_529, dados_limpos_531, dados_limpos_532, dados_limpos_6568, dados_limpos_534  = executando_dados_2(5939)
df_gini = gerando_dataframe5939(dados_limpos_529, dados_limpos_531, dados_limpos_532, dados_limpos_6568, dados_limpos_534)
df_gini.to_excel(ROOT_PATH / 'dados_PIB_gini.xlsx')

#MELHORANDO VISUAL DA PLANILHA
wb_pib = openpyxl.load_workbook(ROOT_PATH / 'dados_PIB_municipais.xlsx')
ws_pib = wb_pib.active
ajustar_bordas(wb_pib)
ajustar_colunas(ws_pib)

wb_pib.save(ROOT_PATH / 'dados_PIB_municipais.xlsx')

wb_pib = openpyxl.load_workbook(ROOT_PATH / 'dados_PIB_estaduais.xlsx')
ws_pib = wb_pib.active
ajustar_bordas(wb_pib)
ajustar_colunas(ws_pib)

wb_pib.save(ROOT_PATH / 'dados_PIB_estaduais.xlsx')

wb_pib = openpyxl.load_workbook(ROOT_PATH / 'dados_PIB_gini.xlsx')
ws_pib = wb_pib.active
ajustar_bordas(wb_pib)
ajustar_colunas(ws_pib)

wb_pib.save(ROOT_PATH / 'dados_PIB_gini.xlsx')

if __name__ == '__main__':
    from sql import executar_sql 
    executar_sql()

